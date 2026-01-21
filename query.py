# -*- coding: utf-8 -*-
import argparse
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
import time
import json
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import os

CITY_LABELS = {
    'hgh': '杭州',
    'sha': '上海',
    'pek': '北京',
    'can': '广州',
    'szx': '深圳',
    'ctu': '成都',
    'akl': '奥克兰',
    'syd': '悉尼',
    'mel': '墨尔本',
}

def log_print(msg):
    """带时间戳的打印函数"""
    timestamp = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    print(f"{timestamp} {msg}")

class CTrip_FlightScraper:
    def __init__(self, headless=True, debug=False):
        # 初始化浏览器
        options = webdriver.ChromeOptions()
        
        if headless:
            # 启用无头模式，不显示浏览器窗口
            options.add_argument('--headless')
        
        # 反爬虫对策
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        options.add_argument('--start-maximized')
        options.add_argument('--disable-extensions')
        options.add_argument('--disable-plugins')
        options.add_argument('--disable-images')  # 禁用图片加载，加快速度
        
        # 伪装成真实浏览器
        options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        options.add_argument('--accept-lang=zh-CN,zh;q=0.9,en;q=0.8')
        
        # 禁用Blink特性识别
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        
        # 使用 webdriver-manager 自动管理 ChromeDriver
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=options)
        self.debug = debug
        
        # 设置隐式等待
        self.driver.implicitly_wait(10)
        
    def scrape_flights(self, url, direct_only=True):
        """
        爬取携程网航班信息
        :param url: 携程航班查询链接
        :return: 航班列表
        """
        try:
            log_print(f"正在访问链接: {url}")
            log_print("请等待，页面加载中...")
            
            # 设置页面加载超时
            self.driver.set_page_load_timeout(30)
            
            self.driver.get(url)
            
            # 先等待任何内容加载
            time.sleep(5)
            
            # 尝试多个等待策略
            try:
                log_print("尝试等待航班项出现...")
                WebDriverWait(self.driver, 15).until(
                    EC.presence_of_all_elements_located((By.CLASS_NAME, "item-inner"))
                )
                log_print("✓ 航班项已加载")
            except:
                try:
                    log_print("尝试等待product元素...")
                    WebDriverWait(self.driver, 10).until(
                        EC.presence_of_all_elements_located((By.CLASS_NAME, "product"))
                    )
                    log_print("✓ product元素已加载")
                except:
                    log_print("⚠ 未能等待到预期的航班元素")
            
            # 额外等待确保动态内容加载完毕
            time.sleep(5)
            
            # 获取页面源代码
            page_source = self.driver.page_source
            
            if self.debug or len(page_source) < 1000:
                # 保存页面源码供调试
                debug_file = "debug_page.html"
                with open(debug_file, 'w', encoding='utf-8') as f:
                    f.write(page_source)
                log_print(f"✓ 页面源码已保存到 {debug_file} (大小: {len(page_source)} 字节)")
            
            if len(page_source) < 100:
                log_print("❌ 获取页面源代码失败，页面过小")
                return []
            
            soup = BeautifulSoup(page_source, 'html.parser')
            
            flights = []
            
            # 使用多个选择器尝试查找航班项
            selectors = [
                ('div', {'class': 'item-inner'}),
                ('div', {'class': 'product'}),
                ('div', {'class': 'flight-item'}),
                ('div', {'class': 'search-item'}),
                ('div', {'class': 'item'}),
            ]
            
            flight_items = []
            for tag, attrs in selectors:
                flight_items = soup.find_all(tag, attrs=attrs)
                if flight_items:
                    log_print(f"✓ 使用选择器 {attrs} 找到 {len(flight_items)} 条记录")
                    break
            
            if not flight_items:
                log_print("❌ 未找到任何航班项")
                # 尝试查看页面中是否有error或提示信息
                error_elem = soup.find('div', class_='error')
                if error_elem:
                    log_print(f"页面提示: {error_elem.get_text()}")
                return []
            
            log_print(f"✓ 找到 {len(flight_items)} 条航班信息")
            
            for idx, item in enumerate(flight_items, 1):
                try:
                    # 过滤直飞航班
                    flight_info = self.parse_flight_item(item, target_direct=direct_only)
                    if flight_info:
                        flights.append(flight_info)
                        log_print(f"  ✓ 成功解析航班 {idx}")
                except Exception as e:
                    if self.debug:
                        log_print(f"  ⚠ 解析航班 {idx} 出错: {e}")
                    continue
            
            return flights
            
        except Exception as e:
            log_print(f"❌ 爬取航班信息失败: {e}")
            import traceback
            traceback.print_exc()
            return []
        finally:
            log_print("正在关闭浏览器...")
            try:
                self.driver.quit()
            except:
                pass
    
    def parse_flight_item(self, item, target_flight_no=None, target_direct=True):
        """
        解析单个航班项，可选过滤目标航班或直飞
        :param item: 航班元素
        :param target_flight_no: 目标航班号
        :param target_direct: 是否仅保留直飞
        """
        try:
            flight_info = {}
            
            # 提取文本与HTML
            item_text = item.get_text(" ", strip=True)
            item_html = str(item)

            # 航班号：抓取所有匹配，直飞只保留单段航班
            flight_no_matches = re.findall(r'([A-Z]{2}\d{2,4})', item_html)
            if not flight_no_matches:
                flight_no_matches = re.findall(r'([A-Z]{2}\d{2,4})', item_text)
            if flight_no_matches:
                flight_info['flight_number'] = flight_no_matches[0]
                # 多个航班号意味着中转/联程
                if target_direct and len(set(flight_no_matches)) > 1:
                    return None

            # 如果指定了目标航班号，则进行过滤
            if target_flight_no and flight_info.get('flight_number'):
                if target_flight_no.upper() not in flight_info['flight_number'].upper():
                    return None

            # 直飞过滤：排除明显含中转/经停的航班
            if target_direct:
                stop_keywords = ['经停', '中转', '转机', '联程', '含中转', '停留']
                if any(k in item_text for k in stop_keywords):
                    return None

            # 航空公司：从文本中抓取“XX航空”或含“航空”的片段
            airline_match = re.search(r'([\u4e00-\u9fa5]{2,6}航空)', item_text)
            if airline_match:
                flight_info['airline'] = airline_match.group(1)

            # 出发/到达时间 (HH:MM)
            time_matches = re.findall(r'(\d{1,2}):(\d{2})', item_text)
            if target_direct and len(time_matches) > 2:
                # 多于两组时间通常是中转
                return None
            if len(time_matches) >= 2:
                flight_info['departure_time'] = f"{time_matches[0][0]}:{time_matches[0][1]}"
                flight_info['arrival_time'] = f"{time_matches[1][0]}:{time_matches[1][1]}"

            # 飞行时长 (xx小时xx分)
            duration_match = re.search(r'(\d+)小时(\d+)分', item_text)
            if duration_match:
                flight_info['duration'] = f"{duration_match.group(1)}h{duration_match.group(2)}m"

            # 价格：匹配合理范围
            price_matches = re.findall(r'¥?\s*(\d+)', item_text)
            if price_matches:
                price_values = []
                for price_str in price_matches:
                    try:
                        price_int = int(price_str)
                        if 1000 <= price_int <= 50000:  # 过滤掉低价误识别（如税费、序号）
                            price_values.append(price_int)
                    except ValueError:
                        continue
                if price_values:
                    flight_info['price'] = str(min(price_values))  # 取最小的合规价格

            # 返回包含价格或航班号的结果
            if flight_info.get('price') or flight_info.get('flight_number'):
                return flight_info
            return None
            
        except Exception as e:
            log_print(f"  ✗ 解析单个航班出错: {e}")
            return None

def save_flights_to_file(flights, filename='flights.json'):
    """
    将航班信息保存到JSON文件
    """
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(flights, f, ensure_ascii=False, indent=2)
    log_print(f"航班信息已保存到 {filename}")


def save_flights_to_excel(flights, dep_city_code, arr_city_code, dep_date, filename='flights_history.xlsx'):
    """
    将航班信息保存到Excel文件（每个航班单独一个sheet：城市对_日期_航空公司_航班号）
    """
    query_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    dep_city_label = city_name(dep_city_code)
    arr_city_label = city_name(arr_city_code)
    
    # 检查文件是否存在
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()
        # 删除默认的Sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
    
    # 为每个航班创建单独的sheet
    for flight in flights:
        airline = flight.get('airline', '未知航空').replace('航空', '')
        flight_no = flight.get('flight_number', 'N/A')
        
        # 生成sheet名称：城市对_日期_航空公司_航班号
        sheet_name = f"{dep_city_label}-{arr_city_label}_{dep_date}_{airline}_{flight_no}"
        
        # Excel sheet名称长度限制为31个字符
        if len(sheet_name) > 31:
            sheet_name = f"{dep_city_label}-{arr_city_label}_{airline}_{flight_no}"[:31]
        
        # 检查sheet是否存在
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            
            # 创建表头
            headers = ['查询时间', '出发城市', '目的地', '出发日期', '航空公司', '航班号', 
                       '出发时间', '到达时间', '飞行时长', '价格(¥)']
            ws.append(headers)
            
            # 设置表头样式
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
        
        # 添加数据行
        row = [
            query_time,
            dep_city_label,
            arr_city_label,
            dep_date,
            flight.get('airline', 'N/A'),
            flight_no,
            flight.get('departure_time', 'N/A'),
            flight.get('arrival_time', 'N/A'),
            flight.get('duration', 'N/A'),
            flight.get('price', 'N/A')
        ]
        ws.append(row)
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filename)
    log_print(f"✓ 航班信息已保存到 {filename}（共 {len(flights)} 个sheet）")

def display_flights(flights, dep_date, dep_city_code="hgh", arr_city_code="akl"):
    """
    显示航班信息 - 直飞航班
    """
    dep_city_label = city_name(dep_city_code)
    arr_city_label = city_name(arr_city_code)
    
    if not flights:
        log_print(f"\n❌ 未找到{dep_city_label} → {arr_city_label}直飞航班")
        log_print("可能原因：")
        log_print("  1. 该日期没有直飞航班")
        log_print("  2. 网站反爬虫保护")
        log_print("  3. 网络连接问题")
        return

    log_print("\n" + "="*80)
    log_print(f"{dep_city_label} → {arr_city_label} 直飞航班（{dep_date}）")
    log_print("="*80)
    log_print(f"{'航空公司':<15} {'航班号':<10} {'出发':<8} {'到达':<8} {'时长':<8} {'价格':<12}")
    log_print("-"*80)
    
    for idx, flight in enumerate(flights, 1):
        log_print(f"{flight.get('airline', 'N/A'):<15} "
              f"{flight.get('flight_number', 'N/A'):<10} "
              f"{flight.get('departure_time', 'N/A'):<8} "
              f"{flight.get('arrival_time', 'N/A'):<8} "
              f"{flight.get('duration', 'N/A'):<8} "
              f"¥ {flight.get('price', 'N/A'):<10}")
    
    log_print("="*80)
    log_print(f"✓ 找到 {len(flights)} 班直飞航班\n")


def build_url(dep_city="hgh", arr_city="akl", dep_date="2026-09-25"):
    return (
        f"https://flights.ctrip.com/online/list/oneway-{dep_city}-{arr_city}?"
        f"depdate={dep_date}&cabin=y_s&adult=1&child=0&infant=0"
    )


def city_name(code: str) -> str:
    return CITY_LABELS.get(code.lower(), code.upper())

# 调用示例：.\.venv\Scripts\python.exe .\query.py --from sha --to akl --date 2026-09-25
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="查询直飞航班")
    parser.add_argument("--from", dest="from_city", default="hgh",
                        help="出发城市代码，如 hgh(杭州)、sha(上海)")
    parser.add_argument("--to", dest="to_city", default="akl",
                        help="到达城市代码，如 akl(奥克兰)、syd(悉尼)")
    parser.add_argument("--date", dest="dep_date", default=datetime.now().strftime("%Y-%m-%d"),
                        help="出发日期，格式YYYY-MM-DD，默认今天")
    parser.add_argument("--headless", dest="headless", action="store_true", default=True,
                        help="启用无头模式，默认开启")
    parser.add_argument("--no-headless", dest="headless", action="store_false",
                        help="关闭无头模式，显示浏览器窗口")
    parser.add_argument("--debug", action="store_true", default=False,
                        help="保存调试页面源码")
    args = parser.parse_args()

    dep_city = args.from_city.strip().lower()
    arr_city = args.to_city.strip().lower()
    
    url = build_url(dep_city=dep_city, arr_city=arr_city, dep_date=args.dep_date)
    dep_label = city_name(dep_city)
    arr_label = city_name(arr_city)
    
    log_print("\n" + "="*80)
    log_print(f"{dep_label} → {arr_label} 直飞航班查询")
    log_print("="*80)
    log_print(f"出发日期: {args.dep_date}")
    log_print(f"出发地: {dep_label} ({dep_city.upper()})")
    log_print(f"目的地: {arr_label} ({arr_city.upper()})")
    log_print("-" * 80)
    
    scraper = CTrip_FlightScraper(headless=args.headless, debug=args.debug)
    flights = scraper.scrape_flights(url, direct_only=True)
    display_flights(flights, dep_date=args.dep_date, dep_city_code=dep_city, arr_city_code=arr_city)
    
    if flights:
        save_flights_to_file(flights, filename=f"flights_{dep_city}_{arr_city}_{args.dep_date}.json")
        save_flights_to_excel(flights, dep_city, arr_city, args.dep_date)
    else:
        log_print("⚠ 未保存任何航班信息")
        log_print("💡 建议: 已保存页面源码到 debug_page.html，请查看页面结构是否改变")

